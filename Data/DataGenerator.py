import openpyxl
import xlrd


def DataGenerator():
    wk = xlrd.open_workbook('../Files/xlrd.xls')
    sh = wk.sheet_by_name('Sheet1')
    r = sh.nrows

    li = []

    for i in range(0, r):
        li1 = []
        un = sh.cell(i, 1)
        up = sh.cell(i, 2)
        ucp = sh.cell(i, 3)
        ue = sh.cell(i, 4)
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        li1.insert(2, ucp.value)
        li1.insert(3, ue.value)
        li.insert(i, li1)

    return li


